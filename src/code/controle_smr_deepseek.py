"""
Contrôle d'exhaustivité SMR — Logique métier
=============================================
Compare les données de passages SMR entre Orbis (agrégé par semaine)
et Hexagone (une ligne par venue) pour identifier les écarts.

Clé de tri : NDA + Date (JJ/MM/AAAA).
La difficulté principale est de transformer les semaines Orbis
(N° semaine + chaîne de Présence) en dates individuelles.
"""

import pandas as pd
import numpy as np
import logging
import argparse
import re
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ==========================================
# CONFIGURATION GLOBALE
# ==========================================
DOSSIER_EXPORT_DEFAUT = Path('./data_test/export_test')

# Colonnes attendues dans chaque fichier source
COLONNES_ORBIS_SMR = ['N° Hospit', 'N° semaine', 'Présence', 'Nom', 'Prénom', 'Né(e) le']
COLONNES_HEXA_SMR = ['Nom/Prénom', 'Nom de Naissance', 'Date de naissance', 'N° Dossier', 'Date', 'Type']


# ==========================================
# FONCTIONS UTILITAIRES
# ==========================================
def valider_colonnes(df, colonnes_attendues, nom_fichier):
    """
    Vérifie que le DataFrame contient bien les colonnes requises.
    """
    df.columns = df.columns.str.strip()
    colonnes_presentes = set(df.columns)
    colonnes_manquantes = [c for c in colonnes_attendues if c not in colonnes_presentes]

    if colonnes_manquantes:
        msg = (f"Le fichier '{nom_fichier}' ne contient pas les colonnes attendues.\n"
               f"  Colonnes manquantes : {colonnes_manquantes}\n"
               f"  Colonnes trouvées   : {list(df.columns)}")
        logging.error(msg)
        raise ValueError(msg)

    logging.info(f"Validation OK pour '{nom_fichier}' ({len(df)} lignes, {len(df.columns)} colonnes)")


def detecter_annee(serie_semaine, nom_fichier=""):
    """
    Détecte intelligemment l'année à utiliser pour les numéros de semaine au format court.
    """
    annees_trouvees = set()
    for val in serie_semaine.dropna().unique():
        val_str = str(val).strip()
        if val_str.endswith('.0'):
            val_str = val_str[:-2]
        if re.match(r'^\d{6}$', val_str):
            annee_candidate = int(val_str[:4])
            if 2000 <= annee_candidate <= 2099:
                annees_trouvees.add(annee_candidate)

    if len(annees_trouvees) == 1:
        annee = annees_trouvees.pop()
        logging.info(f"Année détectée depuis les données Orbis (format long) : {annee}")
        return annee
    elif len(annees_trouvees) > 1:
        annee = max(annees_trouvees)
        logging.warning(f"Plusieurs années détectées dans Orbis : {annees_trouvees}. Utilisation de {annee}.")
        return annee

    # Étape 2 : Chercher une année dans le nom du fichier
    if nom_fichier:
        match = re.search(r'(20\d{2})', nom_fichier)
        if match:
            annee = int(match.group(1))
            logging.info(f"Année détectée depuis le nom du fichier '{nom_fichier}' : {annee}")
            return annee

    # Étape 3 : Fallback → année en cours
    annee = datetime.now().year
    logging.warning(f"Impossible de détecter l'année automatiquement. Utilisation de l'année en cours : {annee}")
    return annee


def parser_semaine_orbis(valeur_semaine, annee_defaut):
    """
    Parse un numéro de semaine Orbis et retourne (année, numéro_semaine).
    """
    val = str(valeur_semaine).strip()
    if val.endswith('.0'):
        val = val[:-2]

    # Format long : 6 chiffres (AAAASS)
    if re.match(r'^\d{6}$', val):
        return int(val[:4]), int(val[4:])

    # Format court : 1 ou 2 chiffres
    if re.match(r'^\d{1,2}$', val):
        return annee_defaut, int(val)

    logging.warning(f"Format de semaine non reconnu : '{val}'")
    return None, None


def eclater_semaine_en_dates(annee, num_semaine, presence_str):
    """
    Transforme une ligne Orbis (semaine + présence) en une liste de datetime.
    """
    dates = []
    presence = str(presence_str).strip()
    # Nettoyer les caractères non autorisés
    presence = re.sub(r'[^LMMJVS\.]', '', presence)
    # Si la chaîne ne fait pas 7 caractères, tenter de corriger
    if len(presence) != 7:
        if len(presence) == 0:
            return dates
        logging.warning(f"Chaîne de présence de longueur {len(presence)} : '{presence_str}' -> '{presence}'")
        if len(presence) < 7:
            presence = presence.ljust(7, '.')
        else:
            presence = presence[:7]
    for idx, char in enumerate(presence):
        if char != '.':
            jour_iso = idx + 1
            try:
                date_obj = datetime.fromisocalendar(annee, num_semaine, jour_iso)
                dates.append(date_obj)
            except ValueError as e:
                logging.warning(f"Date invalide : année={annee}, semaine={num_semaine}, jour={jour_iso} : {e}")
    return dates

def diagnostiquer_erreur(row, df_orbis, df_hexa):
    """
    Diagnostique l'origine de l'erreur en renvoyant 'NDA', 'Semaine' ou 'Jour'.
    """
    nda = str(row['NDA']).strip()
    date_val = row['Date']
    origine = row["Origine de l'écart"]
    nom = str(row['Nom Final'] if 'Nom Final' in row else row.get('Nom', '')).strip().upper()
    prenom = str(row['Prénom Final'] if 'Prénom Final' in row else row.get('Prénom', '')).strip().upper()

    # 1. Erreur de NDA (Faute de frappe)
    if origine == 'Manquant dans Hexagone':
        match_hexa = df_hexa[
            (df_hexa['Nom'].str.strip().str.upper() == nom) &
            (df_hexa['Prénom'].str.strip().str.upper() == prenom) &
            (df_hexa['NDA'] != nda)
        ]
        if not match_hexa.empty:
            return "NDA"
    elif origine == 'Manquant dans Orbis':
        match_orbis = df_orbis[
            (df_orbis['Nom'].str.strip().str.upper() == nom) &
            (df_orbis['Prénom'].str.strip().str.upper() == prenom) &
            (df_orbis['NDA'] != nda)
        ]
        if not match_orbis.empty:
            return "NDA"

    # 2. Erreur de Semaine ou Jour
    nda_dans_orbis = nda in df_orbis['NDA'].values
    nda_dans_hexa = nda in df_hexa['NDA'].values

    try:
        dt_venue = datetime.strptime(date_val, '%d/%m/%Y')
    except Exception:
        dt_venue = None

    if nda_dans_orbis and nda_dans_hexa and dt_venue:
        if origine == 'Manquant dans Hexagone':
            dates_hexa = df_hexa[df_hexa['NDA'] == nda]['Date'].dropna().unique()
            for d_str in dates_hexa:
                try:
                    dt_h = datetime.strptime(d_str, '%d/%m/%Y')
                    diff_days = abs((dt_h - dt_venue).days)
                    if diff_days > 0 and diff_days % 7 == 0:
                        return "Semaine"
                    elif 0 < diff_days < 7:
                        return "Jour"
                except Exception:
                    continue
        elif origine == 'Manquant dans Orbis':
            dates_orbis = df_orbis[df_orbis['NDA'] == nda]['Date'].dropna().unique()
            for d_str in dates_orbis:
                try:
                    dt_o = datetime.strptime(d_str, '%d/%m/%Y')
                    diff_days = abs((dt_venue - dt_o).days)
                    if diff_days > 0 and diff_days % 7 == 0:
                        return "Semaine"
                    elif 0 < diff_days < 7:
                        return "Jour"
                except Exception:
                    continue
                    
    return ""

# ==========================================
# FONCTION PRINCIPALE
# ==========================================
def lancer_controle_smr(orbis_path, hexa_path, export_dir=None):
    """
    Lance le contrôle d'exhaustivité SMR.
    """
    # --- Configuration du dossier d'export ---
    if export_dir:
        dossier_export = Path(export_dir)
    else:
        dossier_export = DOSSIER_EXPORT_DEFAUT
    dossier_export.mkdir(parents=True, exist_ok=True)

    # --- Configuration du fichier de log ---
    date_str = datetime.now().strftime("%Y%m%d")
    log_path = dossier_export / f'Logs_SMR_{date_str}.txt'
    logging.basicConfig(
        filename=log_path,
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        force=True
    )

    logging.info("=== DÉMARRAGE DU CONTRÔLE SMR ===")

    # =========================================================
    # 1. CHARGEMENT ET PRÉPARATION D'ORBIS
    # =========================================================
    logging.info("--- Étape 1 : Chargement Orbis SMR ---")
    chemin_orbis = Path(orbis_path)

    # Lecture sans dtype=str pour éviter les problèmes de conversion
    df_orbis_brut = pd.read_excel(chemin_orbis)
    # Forcer les colonnes textuelles en string
    for col in ['N° Hospit', 'N° semaine', 'Présence', 'Nom', 'Prénom', 'Né(e) le']:
        df_orbis_brut[col] = df_orbis_brut[col].astype(str).str.strip()
    # Supprimer le .0 des semaines
    df_orbis_brut['N° semaine'] = df_orbis_brut['N° semaine'].str.replace(r'\.0$', '', regex=True)

    valider_colonnes(df_orbis_brut, COLONNES_ORBIS_SMR, chemin_orbis.name)
    nb_orbis_brut = len(df_orbis_brut)
    logging.info(f"Orbis brut : {nb_orbis_brut} lignes")

    # --- Détection de l'année ---
    annee_defaut = detecter_annee(df_orbis_brut['N° semaine'], chemin_orbis.name)
    logging.info(f"Année par défaut pour les semaines courtes : {annee_defaut}")

    # --- Éclatement des semaines en dates ---
    lignes_eclatees = []
    lignes_semaine_invalide = 0
    lignes_presence_invalide = 0

    for _, row in df_orbis_brut.iterrows():
        annee, num_sem = parser_semaine_orbis(row['N° semaine'], annee_defaut)
        if annee is None or num_sem is None:
            lignes_semaine_invalide += 1
            logging.warning(f"Semaine invalide : {row['N° semaine']} pour NDA {row['N° Hospit']}")
            continue
        dates_obj = eclater_semaine_en_dates(annee, num_sem, row['Présence'])
        if not dates_obj:
            lignes_presence_invalide += 1
            logging.warning(f"Aucune date générée pour NDA {row['N° Hospit']}, semaine {row['N° semaine']}, présence '{row['Présence']}'")
        for date_obj in dates_obj:
            lignes_eclatees.append({
                'NDA': row['N° Hospit'],
                'Date': date_obj,  # datetime
                'Nom': row['Nom'],
                'Prénom': row['Prénom'],
                'Né(e) le': row['Né(e) le']
            })

    df_orbis = pd.DataFrame(lignes_eclatees)
    nb_orbis_eclate = len(df_orbis)
    logging.info(f"Orbis après éclatement : {nb_orbis_eclate} lignes (venues individuelles)")
    logging.info(f"Lignes Orbis ignorées : {lignes_semaine_invalide} (semaine invalide), {lignes_presence_invalide} (présence sans date)")

    # =========================================================
    # 2. CHARGEMENT ET PRÉPARATION D'HEXAGONE
    # =========================================================
    logging.info("--- Étape 2 : Chargement Hexagone SMR ---")
    chemin_hexa = Path(hexa_path)

    # Lecture sans dtype=str, avec header=2
    df_hexa_brut = pd.read_excel(chemin_hexa, header=2)
    # Convertir la colonne Date avec le format spécifique
    df_hexa_brut['Date'] = pd.to_datetime(
        df_hexa_brut['Date'],
        format='%d/%m/%Y %H:%M:%S',
        errors='coerce'
    )

    valider_colonnes(df_hexa_brut, COLONNES_HEXA_SMR, chemin_hexa.name)
    nb_hexa_brut = len(df_hexa_brut)
    logging.info(f"Hexagone brut : {nb_hexa_brut} lignes")

    # Suppression des lignes SD
    nb_sd = (df_hexa_brut['Type'].str.strip() == 'SD').sum()
    df_hexa = df_hexa_brut[df_hexa_brut['Type'].str.strip() != 'SD'].copy()
    nb_hexa_apres_sd = len(df_hexa)
    logging.info(f"Lignes SD supprimées : {nb_sd}. Hexagone après nettoyage : {nb_hexa_apres_sd} lignes")

    # --- Nettoyage des colonnes ---
    df_hexa['NDA'] = df_hexa['N° Dossier'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    # Extraire Nom et Prénom
    split_noms = df_hexa['Nom/Prénom'].str.split('/', n=1, expand=True)
    df_hexa['Nom'] = split_noms[0]
    df_hexa['Prénom'] = split_noms[1] if 1 in split_noms.columns else ''

    # Créer une colonne Date sans heure pour le merge
    df_hexa['Date_juste'] = df_hexa['Date'].dt.normalize()

    nb_dates_ok = df_hexa['Date_juste'].notna().sum()
    logging.info(f"Hexagone dates parsées : {nb_dates_ok}/{nb_hexa_apres_sd}")

    # =========================================================
    # 3. TRI : COMPARAISON SUR CLÉ COMPOSITE NDA + DATE
    # =========================================================
    logging.info("--- Étape 3 : Tri NDA + Date ---")

    # Colonnes à garder pour le merge
    cols_orbis = ['NDA', 'Date', 'Nom', 'Prénom', 'Né(e) le']
    cols_hexa = ['NDA', 'Date_juste', 'Nom', 'Prénom', 'Date de naissance']

    # Dédoublonnage
    orbis_dedup = df_orbis[cols_orbis].drop_duplicates(subset=['NDA', 'Date'])
    hexa_dedup = df_hexa[cols_hexa].drop_duplicates(subset=['NDA', 'Date_juste']).rename(columns={'Date_juste': 'Date'})

    logging.info(f"Orbis pour le merge : {len(orbis_dedup)} lignes")
    logging.info(f"Hexagone pour le merge : {len(hexa_dedup)} lignes")

    # Jointure externe
    tri = pd.merge(
        orbis_dedup,
        hexa_dedup,
        on=['NDA', 'Date'],
        how='outer',
        indicator=True,
        suffixes=('_orbis', '_hexa')
    )

    anomalies = tri[tri['_merge'] != 'both'].copy()
    nb_match = (tri['_merge'] == 'both').sum()
    nb_anomalies = len(anomalies)
    logging.info(f"Résultat du tri — Correspondances : {nb_match}, Anomalies : {nb_anomalies}")

    # --- Colonne d'origine de l'écart ---
    if not anomalies.empty:
        anomalies['Origine de l\'écart'] = anomalies['_merge'].map({
            'left_only': 'Manquant dans Hexagone',
            'right_only': 'Manquant dans Orbis'
        })

        # Consolidation des colonnes Nom/Prénom/Date naissance
        anomalies['Nom Final'] = anomalies['Nom_orbis'].combine_first(anomalies['Nom_hexa'])
        anomalies['Prénom Final'] = anomalies['Prénom_orbis'].combine_first(anomalies['Prénom_hexa'])
        anomalies['Date Naissance Final'] = anomalies['Né(e) le'].combine_first(anomalies['Date de naissance'])

    # --- Préparation du DataFrame d'export ---
    colonnes_export = ['NDA', 'Nom Final', 'Prénom Final', 'Date Naissance Final', 'Date', 'Origine de l\'écart']
    if not anomalies.empty:
        # Diagnostic d'erreur (appel à une fonction externe, la vôtre)
        # Si vous avez la fonction diagnostiquer_erreur, réutilisez-la ici
        # Sinon, la colonne "Origine de l'erreur" reste vide
        anomalies["Origine de l'erreur"] = ""  # ou appelez votre fonction
        colonnes_export.append("Origine de l'erreur")

        export_tri = anomalies[colonnes_export].rename(columns={
            'Nom Final': 'Nom',
            'Prénom Final': 'Prénom',
            'Date Naissance Final': 'Date Naissance',
        })

        # Tri par date décroissante
        export_tri = export_tri.sort_values('Date', ascending=False)
    else:
        export_tri = pd.DataFrame(columns=[
            'NDA', 'Nom', 'Prénom', 'Date Naissance', 'Date', "Origine de l'écart", "Origine de l'erreur"
        ])

    # Compteurs pour la synthèse
    nb_manque_hexa = len(export_tri[export_tri["Origine de l'écart"] == 'Manquant dans Hexagone'])
    nb_manque_orbis = len(export_tri[export_tri["Origine de l'écart"] == 'Manquant dans Orbis'])

    # =========================================================
    # 4. EXPORT VERS EXCEL
    # =========================================================
    logging.info("--- Étape 4 : Export des résultats ---")

    # Fichier de tri (anomalies)
    chemin_tri = dossier_export / f'Tri_SMR_Ecarts_{date_str}.xlsx'
    export_tri.to_excel(chemin_tri, index=False, engine='openpyxl')
    mettre_en_forme_excel(chemin_tri)
    logging.info(f"Fichier de tri exporté : {chemin_tri.name} ({len(export_tri)} lignes)")

    # Fichier de synthèse
    chemin_synthese = dossier_export / f'Synthese_SMR_{date_str}.xlsx'
    synthese_data = {
        'Indicateur': [
            'Date du traitement',
            '---',
            'DONNEES EN ENTREE',
            'Lignes Orbis (brut, par semaine)',
            f'Année détectée pour semaines courtes',
            'Lignes Orbis (après éclatement en dates)',
            'Lignes Hexagone (brut)',
            'Lignes SD supprimées',
            'Lignes Hexagone (après nettoyage)',
            '---',
            'TRI SMR - ECARTS NDA + DATE',
            'Correspondances trouvées',
            'Anomalies totales',
            'Manquant dans Hexagone',
            'Manquant dans Orbis',
        ],
        'Valeur': [
            datetime.now().strftime("%d/%m/%Y %H:%M"),
            '',
            '',
            nb_orbis_brut,
            annee_defaut,
            nb_orbis_eclate,
            nb_hexa_brut,
            nb_sd,
            nb_hexa_apres_sd,
            '',
            '',
            nb_match,
            nb_anomalies,
            nb_manque_hexa,
            nb_manque_orbis,
        ]
    }
    df_synthese = pd.DataFrame(synthese_data)
    df_synthese.to_excel(chemin_synthese, index=False, engine='openpyxl')
    mettre_en_forme_excel(chemin_synthese)
    logging.info(f"Fichier de synthèse exporté : {chemin_synthese.name}")

    logging.info("=== CONTRÔLE SMR TERMINÉ ===")


def mettre_en_forme_excel(chemin_fichier):
    """
    Applique la mise en forme sur un fichier Excel existant.
    """
    wb = load_workbook(chemin_fichier)

    style_entete = Font(bold=True, color="FFFFFF", size=11)
    fond_entete = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    alignement = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ws in wb.worksheets:
        if ws.max_row < 1:
            continue

        # En-têtes
        for cell in ws[1]:
            cell.font = style_entete
            cell.fill = fond_entete
            cell.alignment = alignement

        # Auto-ajustement
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

        # Filtres
        ws.auto_filter.ref = ws.dimensions

        # Format date pour les colonnes contenant des datetime
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell.number_format = 'DD/MM/YYYY'

    wb.save(chemin_fichier)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Contrôle d'exhaustivité SMR — DIM")
    parser.add_argument('--orbis', required=True, help='Chemin du fichier Orbis SMR')
    parser.add_argument('--hexa', required=True, help='Chemin du fichier Hexagone SMR')
    parser.add_argument('--export', help="Dossier d'export (défaut: data_test/export_test)")
    args = parser.parse_args()

    lancer_controle_smr(
        orbis_path=args.orbis,
        hexa_path=args.hexa,
        export_dir=args.export
    )