"""
Contrôle d'exhaustivité SMR — Logique métier
=============================================
Compare les données de passages SMR entre Orbis (agrégé par semaine)
et Hexagone (une ligne par venue) pour identifier les écarts.

Clé de tri : NDA + Date (JJ/MM/AAAA).
La difficulté principale est de transformer les semaines Orbis
(N° semaine + chaîne de Présence) en dates individuelles.
"""

import pandas as pd
import numpy as np
import logging
import argparse
import re
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ==========================================
# CONFIGURATION GLOBALE
# ==========================================
DOSSIER_EXPORT_DEFAUT = Path('./data_test/export_test')

# Colonnes attendues dans chaque fichier source
COLONNES_ORBIS_SMR = ['N° Hospit', 'N° semaine', 'Présence', 'Nom', 'Prénom', 'Né(e) le']
COLONNES_HEXA_SMR = ['Nom/Prénom', 'Nom de Naissance', 'Date de naissance', 'N° Dossier', 'Date', 'Type']


# ==========================================
# FONCTIONS UTILITAIRES
# ==========================================
def valider_colonnes(df, colonnes_attendues, nom_fichier):
    """
    Vérifie que le DataFrame contient bien les colonnes requises.
    Si le format change (mise à jour du logiciel source), on plante proprement 
    au lieu d'avoir des erreurs illisibles plus bas.
    """
    df.columns = df.columns.str.strip()
    colonnes_presentes = set(df.columns)
    colonnes_manquantes = [c for c in colonnes_attendues if c not in colonnes_presentes]

    if colonnes_manquantes:
        msg = (f"Le fichier '{nom_fichier}' ne contient pas les colonnes attendues.\n"
               f"  Colonnes manquantes : {colonnes_manquantes}\n"
               f"  Colonnes trouvées   : {list(df.columns)}")
        logging.error(msg)
        raise ValueError(msg)

    logging.info(f"Validation OK pour '{nom_fichier}' ({len(df)} lignes, {len(df.columns)} colonnes)")


def detecter_annee(serie_semaine, nom_fichier=""):
    """
    Détecte intelligemment l'année à utiliser pour les numéros de semaine au format court.
    
    Stratégie en 3 étapes :
    1. Cherche dans les autres lignes Orbis un format long (ex: '202608') pour en extraire l'année.
    2. Cherche une année à 4 chiffres dans le nom du fichier source.
    3. Fallback : utilise l'année en cours du PC.
    
    Retourne l'année (int) détectée.
    """
    # Étape 1 : Scanner les valeurs au format long (AAAASS = 6 chiffres)
    annees_trouvees = set()
    for val in serie_semaine.dropna().unique():
        val_str = str(val).strip()
        # Supprime le '.0' résiduel de la lecture Excel (ex: '202603.0' → '202603')
        if val_str.endswith('.0'):
            val_str = val_str[:-2]
        # Format long : 6 chiffres (ex: '202608')
        if re.match(r'^\d{6}$', val_str):
            annee_candidate = int(val_str[:4])
            # Filtre de vraisemblance (entre 2000 et 2099)
            if 2000 <= annee_candidate <= 2099:
                annees_trouvees.add(annee_candidate)

    if len(annees_trouvees) == 1:
        annee = annees_trouvees.pop()
        logging.info(f"Année détectée depuis les données Orbis (format long) : {annee}")
        return annee
    elif len(annees_trouvees) > 1:
        # Plusieurs années trouvées → on prend la plus fréquente
        annee = max(annees_trouvees)
        logging.warning(f"Plusieurs années détectées dans Orbis : {annees_trouvees}. Utilisation de {annee}.")
        return annee

    # Étape 2 : Chercher une année dans le nom du fichier
    if nom_fichier:
        match = re.search(r'(20\d{2})', nom_fichier)
        if match:
            annee = int(match.group(1))
            logging.info(f"Année détectée depuis le nom du fichier '{nom_fichier}' : {annee}")
            return annee

    # Étape 3 : Fallback → année en cours
    annee = datetime.now().year
    logging.warning(f"Impossible de détecter l'année automatiquement. Utilisation de l'année en cours : {annee}")
    return annee


def parser_semaine_orbis(valeur_semaine, annee_defaut):
    """
    Parse un numéro de semaine Orbis et retourne (année, numéro_semaine).
    
    Formats gérés :
    - '202608' → (2026, 8)
    - '8'      → (annee_defaut, 8)
    - '02'     → (annee_defaut, 2)
    """
    val = str(valeur_semaine).strip()
    # Supprime le '.0' résiduel de la lecture Excel (ex: '202603.0' → '202603')
    if val.endswith('.0'):
        val = val[:-2]

    # Format long : 6 chiffres (AAAASS)
    if re.match(r'^\d{6}$', val):
        return int(val[:4]), int(val[4:])

    # Format court : 1 ou 2 chiffres (numéro de semaine seul)
    if re.match(r'^\d{1,2}$', val):
        return annee_defaut, int(val)

    # Valeur inattendue → on retourne None pour la gérer plus tard
    logging.warning(f"Format de semaine non reconnu : '{val}'")
    return None, None


def eclater_semaine_en_dates(annee, num_semaine, presence_str):
    """
    Transforme une ligne Orbis (semaine + présence) en une liste de dates JJ/MM/AAAA.
    
    Chaque caractère de la chaîne de présence correspond à un jour :
    - Position 0 = Lundi   (L)
    - Position 1 = Mardi   (M)
    - Position 2 = Mercredi(M)
    - Position 3 = Jeudi   (J)
    - Position 4 = Vendredi(V)
    - Position 5 = Samedi  (S)
    - Position 6 = Dimanche(D)
    Un point '.' signifie absence, une lettre signifie présence.
    
    Utilise datetime.fromisocalendar() pour calculer la date exacte
    à partir de (année, semaine_iso, jour_semaine).
    
    Retourne une liste de strings au format JJ/MM/AAAA.
    """
    dates = []
    presence = str(presence_str).strip()

    # Nettoyer les caractères non autorisés
    presence = re.sub(r'[^LMMJVS\.]', '', presence)
    
    # On s'assure que la chaîne fait bien 7 caractères
    if len(presence) != 7:
        logging.warning(f"Chaîne de présence invalide (longueur {len(presence)}) : '{presence}'")
        return dates

    for idx, char in enumerate(presence):
        if char != '.':
            # idx 0 = Lundi (jour ISO 1), idx 6 = Dimanche (jour ISO 7)
            jour_iso = idx + 1
            try:
                date_obj = datetime.fromisocalendar(annee, num_semaine, jour_iso)
                dates.append(date_obj)
            except ValueError as e:
                logging.warning(f"Date invalide pour année={annee}, semaine={num_semaine}, jour={jour_iso}: {e}")

    return dates


def formater_date_hexa(date_val):
    """
    Convertit une date Hexagone en objet datetime (sans heure).
    
    Gère 3 cas selon comment Excel/Pandas a lu la cellule :
    1. datetime natif (format Personnalisé Excel → openpyxl retourne un datetime)
    2. Chaîne "06/01/2026 09:12:00" (si cellule en format Texte dans Excel)
    3. Chaîne ISO "2026-01-06 09:12:00" (si dtype=str a converti un datetime natif)
    
    Retourne un objet datetime (à minuit) ou pd.NaT si invalide.
    """
    if pd.isna(date_val):
        return pd.NaT
    
    # Cas 1 : Déjà un datetime Python (natif Excel lu par openpyxl)
    if isinstance(date_val, datetime):
        return datetime(date_val.year, date_val.month, date_val.day)
    
    s = str(date_val).strip()
    if s in ('', 'nan', 'NaN', 'NaT', 'None'):
        return pd.NaT

    try:
        # dayfirst=True force l'interprétation JJ/MM/AAAA
        # (fonctionne aussi sur le format ISO qui est non ambigu)
        dt_val = pd.to_datetime(s, dayfirst=True)
        return datetime(dt_val.year, dt_val.month, dt_val.day)
    except Exception:
        logging.warning(f"Date Hexagone non parsable : '{date_val}'")
        return pd.NaT


def mettre_en_forme_excel(chemin_fichier):
    """
    Applique la mise en forme sur un fichier Excel existant :
    - En-têtes en gras avec fond bleu
    - Largeur de colonnes auto-ajustée
    - Filtres automatiques activés
    Identique à la mission 1 pour garder une cohérence visuelle.
    """
    wb = load_workbook(chemin_fichier)

    style_entete = Font(bold=True, color="FFFFFF", size=11)
    fond_entete = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    alignement = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ws in wb.worksheets:
        if ws.max_row < 1:
            continue

        # Mise en forme des en-têtes (ligne 1)
        for cell in ws[1]:
            cell.font = style_entete
            cell.fill = fond_entete
            cell.alignment = alignement

        # Auto-ajustement de la largeur des colonnes
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

        # Activation des filtres automatiques
        ws.auto_filter.ref = ws.dimensions

        # Formatage explicite des cellules datetime au format JJ/MM/AAAA
        # (évite qu'Excel reformate les dates selon la locale du poste)
        for col_idx in range(1, ws.max_column + 1):
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, datetime):
                    cell.number_format = 'DD/MM/YYYY'

    wb.save(chemin_fichier)


def diagnostiquer_erreur(row, df_orbis, df_hexa):
    """
    Diagnostique l'origine de l'erreur en renvoyant 'NDA', 'Semaine' ou 'Jour'.
    """
    nda = str(row['NDA']).strip()
    date_val = row['Date']
    origine = row["Origine de l'écart"]
    nom = str(row['Nom Final'] if 'Nom Final' in row else row.get('Nom', '')).strip().upper()
    prenom = str(row['Prénom Final'] if 'Prénom Final' in row else row.get('Prénom', '')).strip().upper()

    # 1. Erreur de NDA (Faute de frappe)
    if origine == 'Manquant dans Hexagone':
        match_hexa = df_hexa[
            (df_hexa['Nom'].str.strip().str.upper() == nom) &
            (df_hexa['Prénom'].str.strip().str.upper() == prenom) &
            (df_hexa['NDA'] != nda)
        ]
        if not match_hexa.empty:
            return "NDA"
    elif origine == 'Manquant dans Orbis':
        match_orbis = df_orbis[
            (df_orbis['Nom'].str.strip().str.upper() == nom) &
            (df_orbis['Prénom'].str.strip().str.upper() == prenom) &
            (df_orbis['NDA'] != nda)
        ]
        if not match_orbis.empty:
            return "NDA"

    # 2. Erreur de Semaine ou Jour
    nda_dans_orbis = nda in df_orbis['NDA'].values
    nda_dans_hexa = nda in df_hexa['NDA'].values

    # Conversion de la date de la venue (datetime/Timestamp depuis le merge)
    if isinstance(date_val, datetime):
        dt_venue = date_val
    else:
        try:
            dt_venue = pd.to_datetime(date_val, dayfirst=True).to_pydatetime()
        except Exception:
            dt_venue = None

    if nda_dans_orbis and nda_dans_hexa and dt_venue:
        if origine == 'Manquant dans Hexagone':
            dates_hexa = df_hexa[df_hexa['NDA'] == nda]['Date'].dropna().unique()
            for d in dates_hexa:
                try:
                    dt_h = pd.Timestamp(d).to_pydatetime()
                    diff_days = abs((dt_h - dt_venue).days)
                    if diff_days > 0 and diff_days % 7 == 0:
                        return "Semaine"
                    elif 0 < diff_days < 7:
                        return "Jour"
                except Exception:
                    continue
        elif origine == 'Manquant dans Orbis':
            dates_orbis = df_orbis[df_orbis['NDA'] == nda]['Date'].dropna().unique()
            for d in dates_orbis:
                try:
                    dt_o = pd.Timestamp(d).to_pydatetime()
                    diff_days = abs((dt_venue - dt_o).days)
                    if diff_days > 0 and diff_days % 7 == 0:
                        return "Semaine"
                    elif 0 < diff_days < 7:
                        return "Jour"
                except Exception:
                    continue
                    
    return ""

# ==========================================
# FONCTION PRINCIPALE
# ==========================================
def lancer_controle_smr(orbis_path, hexa_path, export_dir=None):
    """
    Lance le contrôle d'exhaustivité SMR.
    
    Paramètres :
    - orbis_path  : chemin vers le fichier Orbis SMR (.xls/.xlsx)
    - hexa_path   : chemin vers le fichier Hexagone SMR (.xls/.xlsx)
    - export_dir  : dossier de sortie (optionnel, défaut = data_test/export_test)
    """
    # --- Configuration du dossier d'export ---
    if export_dir:
        dossier_export = Path(export_dir)
    else:
        dossier_export = DOSSIER_EXPORT_DEFAUT
    dossier_export.mkdir(parents=True, exist_ok=True)

    # --- Configuration du fichier de log ---
    # force=True permet d'écraser la configuration si on relance depuis l'interface
    date_str = datetime.now().strftime("%Y%m%d")
    log_path = dossier_export / f'Logs_SMR_{date_str}.txt'
    logging.basicConfig(
        filename=log_path,
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        force=True
    )

    logging.info("=== DÉMARRAGE DU CONTRÔLE SMR ===")

    # =========================================================
    # 1. CHARGEMENT ET PRÉPARATION D'ORBIS
    # =========================================================
    logging.info("--- Étape 1 : Chargement Orbis SMR ---")
    chemin_orbis = Path(orbis_path)
    
    # Lecture native (sans dtype=str) pour préserver les types Excel (dates en datetime)
    df_orbis_brut = pd.read_excel(chemin_orbis)
    valider_colonnes(df_orbis_brut, COLONNES_ORBIS_SMR, chemin_orbis.name)

    # Forcer les colonnes textuelles en string après lecture native
    # On laisse 'Né(e) le' en type natif (datetime) pour conserver les dates de naissance
    for col in ['N° Hospit', 'N° semaine', 'Présence', 'Nom', 'Prénom']:
        df_orbis_brut[col] = df_orbis_brut[col].astype(str).str.strip()
    # Supprime le '.0' résiduel des colonnes numériques lues en float par Excel
    for col in ['N° Hospit', 'N° semaine']:
        df_orbis_brut[col] = df_orbis_brut[col].str.replace(r'\.0$', '', regex=True)
    
    nb_orbis_brut = len(df_orbis_brut)
    logging.info(f"Orbis brut : {nb_orbis_brut} lignes")

    # --- Détection intelligente de l'année ---
    annee_defaut = detecter_annee(df_orbis_brut['N° semaine'], chemin_orbis.name)
    logging.info(f"Année par défaut pour les semaines courtes : {annee_defaut}")

    # --- Éclatement des semaines en dates individuelles ---
    # Chaque ligne Orbis (1 semaine, ex: 'L.M.V..') va produire N lignes (une par jour de présence)
    lignes_eclatees = []
    for _, row in df_orbis_brut.iterrows():
        annee, num_sem = parser_semaine_orbis(row['N° semaine'], annee_defaut)
        if annee is None or num_sem is None:
            logging.warning(f"Ligne Orbis ignorée (semaine invalide) : NDA={row['N° Hospit']}, Semaine={row['N° semaine']}")
            continue

        dates = eclater_semaine_en_dates(annee, num_sem, row['Présence'])
        for date_venue in dates:
            lignes_eclatees.append({
                'NDA': str(row['N° Hospit']).strip(),
                'Date': date_venue,
                'Nom': row['Nom'],
                'Prénom': row['Prénom'],
                'Né(e) le': row['Né(e) le'],
                'N° semaine (source)': row['N° semaine'],
                'Présence (source)': row['Présence'],
            })

    df_orbis = pd.DataFrame(lignes_eclatees)
    nb_orbis_eclate = len(df_orbis)
    # Normalisation en datetime64 pandas à minuit pour le merge
    if not df_orbis.empty:
        df_orbis['Date'] = pd.to_datetime(df_orbis['Date'], errors='coerce').dt.normalize()
    logging.info(f"Orbis après éclatement : {nb_orbis_eclate} lignes (venues individuelles)")

    # =========================================================
    # 2. CHARGEMENT ET PRÉPARATION D'HEXAGONE
    # =========================================================
    logging.info("--- Étape 2 : Chargement Hexagone SMR ---")
    chemin_hexa = Path(hexa_path)

    # Les fichiers Hexagone ont 2 lignes d'en-tête (titre + ligne vide) avant les colonnes
    # Lecture native (sans dtype=str) pour préserver les dates Excel en datetime
    df_hexa_brut = pd.read_excel(chemin_hexa, header=2)
    valider_colonnes(df_hexa_brut, COLONNES_HEXA_SMR, chemin_hexa.name)

    nb_hexa_brut = len(df_hexa_brut)
    logging.info(f"Hexagone brut : {nb_hexa_brut} lignes")

    # --- Suppression des lignes SD ---
    # Les lignes de type 'SD' (Sortie de Domicile) font doublon avec la venue précédente 
    # (même jour, juste décalée de quelques heures). On les supprime avant le tri.
    df_hexa_brut['Type'] = df_hexa_brut['Type'].astype(str).str.strip()
    nb_sd = (df_hexa_brut['Type'] == 'SD').sum()
    df_hexa = df_hexa_brut[df_hexa_brut['Type'] != 'SD'].copy()
    nb_hexa_apres_sd = len(df_hexa)
    logging.info(f"Lignes SD supprimées : {nb_sd}. Hexagone après nettoyage : {nb_hexa_apres_sd} lignes")

    # --- Nettoyage des colonnes ---
    df_hexa['NDA'] = df_hexa['N° Dossier'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    # Normalisation des dates : conversion en datetime64 à minuit (suppression de l'heure)
    # Les dates Excel natives sont des datetime, certaines cellules texte sont parsées avec dayfirst=True
    df_hexa['Date'] = pd.to_datetime(df_hexa['Date'], dayfirst=True, errors='coerce').dt.normalize()

    # Extraction du Nom et du Prénom depuis 'Nom/Prénom'
    split_noms = df_hexa['Nom/Prénom'].str.split('/', n=1, expand=True)
    df_hexa['Nom'] = split_noms[0]
    df_hexa['Prénom'] = split_noms[1] if 1 in split_noms.columns else ''

    nb_dates_ok = df_hexa['Date'].notna().sum()
    logging.info(f"Hexagone dates parsées : {nb_dates_ok}/{nb_hexa_apres_sd}")

    # =========================================================
    # 3. TRI : COMPARAISON SUR CLÉ COMPOSITE NDA + DATE
    # =========================================================
    logging.info("--- Étape 3 : Tri NDA + Date ---")

    # Colonnes à garder pour le merge
    cols_orbis = ['NDA', 'Date', 'Nom', 'Prénom', 'Né(e) le']
    cols_hexa = ['NDA', 'Date', 'Nom', 'Prénom', 'Date de naissance']

    # Dédoublonnage avant jointure pour éviter les produits cartésiens
    # (un même patient peut avoir des doublons si 2 visites le même jour dans Hexagone)
    orbis_dedup = df_orbis[cols_orbis].drop_duplicates(subset=['NDA', 'Date'])
    hexa_dedup = df_hexa[cols_hexa].drop_duplicates(subset=['NDA', 'Date'])

    logging.info(f"Orbis pour le merge : {len(orbis_dedup)} lignes (dédoublonnées)")
    logging.info(f"Hexagone pour le merge : {len(hexa_dedup)} lignes (dédoublonnées)")

    # --- Jointure Externe (Outer Join) ---
    # On joint sur la clé composite (NDA, Date).
    # indicator=True ajoute une colonne '_merge' pour savoir d'où vient chaque ligne.
    tri = pd.merge(
        orbis_dedup,
        hexa_dedup,
        on=['NDA', 'Date'],
        how='outer',
        indicator=True,
        suffixes=('_orbis', '_hexa')
    )

    # Les anomalies sont les lignes qui ne sont PAS dans 'both'
    anomalies = tri[tri['_merge'] != 'both'].copy()
    nb_match = (tri['_merge'] == 'both').sum()
    nb_anomalies = len(anomalies)
    logging.info(f"Résultat du tri — Correspondances : {nb_match}, Anomalies : {nb_anomalies}")

    # --- Colonne d'origine de l'écart ---
    if not anomalies.empty:
        anomalies['Origine de l\'écart'] = anomalies['_merge'].map({
            'left_only': 'Manquant dans Hexagone',
            'right_only': 'Manquant dans Orbis'
        })

        # Consolidation des colonnes Nom/Prénom/Date naissance (prendre la valeur dispo)
        anomalies['Nom Final'] = anomalies['Nom_orbis'].combine_first(anomalies['Nom_hexa'])
        anomalies['Prénom Final'] = anomalies['Prénom_orbis'].combine_first(anomalies['Prénom_hexa'])
        anomalies['Date Naissance Final'] = anomalies['Né(e) le'].combine_first(anomalies['Date de naissance'])

    # --- Préparation du DataFrame d'export ---
    colonnes_export = ['NDA', 'Nom Final', 'Prénom Final', 'Date Naissance Final', 'Date', 'Origine de l\'écart']
    if not anomalies.empty:
        # Calcul du diagnostic d'erreur
        anomalies["Origine de l'erreur"] = anomalies.apply(lambda r: diagnostiquer_erreur(r, df_orbis, df_hexa), axis=1)
        colonnes_export.append("Origine de l'erreur")
        
        export_tri = anomalies[colonnes_export].rename(columns={
            'Nom Final': 'Nom',
            'Prénom Final': 'Prénom',
            'Date Naissance Final': 'Date Naissance',
        })

        # Les dates sont déjà des objets datetime depuis les 2 sources (Orbis et Hexagone)
        # On normalise en datetime64 pandas pour un export Excel propre en JJ/MM/AAAA
        export_tri['Date'] = pd.to_datetime(export_tri['Date'], errors='coerce')
        if 'Date Naissance' in export_tri.columns:
            export_tri['Date Naissance'] = pd.to_datetime(export_tri['Date Naissance'], dayfirst=True, errors='coerce')
        # Tri par date décroissante pour faciliter la lecture
        export_tri = export_tri.sort_values('Date', ascending=False)
    else:
        export_tri = pd.DataFrame(columns=[
            'NDA', 'Nom', 'Prénom', 'Date Naissance', 'Date', "Origine de l'écart", "Origine de l'erreur"
        ])

    # Compteurs pour la synthèse
    nb_manque_hexa = len(export_tri[export_tri["Origine de l'écart"] == 'Manquant dans Hexagone'])
    nb_manque_orbis = len(export_tri[export_tri["Origine de l'écart"] == 'Manquant dans Orbis'])

    # =========================================================
    # 4. EXPORT VERS EXCEL
    # =========================================================
    logging.info("--- Étape 4 : Export des résultats ---")

    # Fichier de tri (anomalies)
    chemin_tri = dossier_export / f'Tri_SMR_Ecarts_{date_str}.xlsx'
    export_tri.to_excel(chemin_tri, index=False, engine='openpyxl')
    mettre_en_forme_excel(chemin_tri)
    logging.info(f"Fichier de tri exporté : {chemin_tri.name} ({len(export_tri)} lignes)")

    # Fichier de synthèse
    chemin_synthese = dossier_export / f'Synthese_SMR_{date_str}.xlsx'
    synthese_data = {
        'Indicateur': [
            'Date du traitement',
            '---',
            'DONNEES EN ENTREE',
            'Lignes Orbis (brut, par semaine)',
            f'Année détectée pour semaines courtes',
            'Lignes Orbis (après éclatement en dates)',
            'Lignes Hexagone (brut)',
            'Lignes SD supprimées',
            'Lignes Hexagone (après nettoyage)',
            '---',
            'TRI SMR - ECARTS NDA + DATE',
            'Correspondances trouvées',
            'Anomalies totales',
            'Manquant dans Hexagone',
            'Manquant dans Orbis',
        ],
        'Valeur': [
            datetime.now().strftime("%d/%m/%Y %H:%M"),
            '',
            '',
            nb_orbis_brut,
            annee_defaut,
            nb_orbis_eclate,
            nb_hexa_brut,
            nb_sd,
            nb_hexa_apres_sd,
            '',
            '',
            nb_match,
            nb_anomalies,
            nb_manque_hexa,
            nb_manque_orbis,
        ]
    }
    df_synthese = pd.DataFrame(synthese_data)
    df_synthese.to_excel(chemin_synthese, index=False, engine='openpyxl')
    mettre_en_forme_excel(chemin_synthese)
    logging.info(f"Fichier de synthèse exporté : {chemin_synthese.name}")

    logging.info("=== CONTRÔLE SMR TERMINÉ ===")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Contrôle d'exhaustivité SMR — DIM")
    parser.add_argument('--orbis', required=True, help='Chemin du fichier Orbis SMR')
    parser.add_argument('--hexa', required=True, help='Chemin du fichier Hexagone SMR')
    parser.add_argument('--export', help="Dossier d'export (défaut: data_test/export_test)")
    args = parser.parse_args()

    lancer_controle_smr(
        orbis_path=args.orbis,
        hexa_path=args.hexa,
        export_dir=args.export
    )
